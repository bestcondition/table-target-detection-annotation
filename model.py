from collections import namedtuple
from pathlib import Path
import numpy as np

import openpyxl

"""
y是行号，从零开始
x是列号，从零开始
h是高，有多少行
w是宽，有多少列
mark是标注的类名

tuple 直接可哈希，因为不可能有这五个值完全相同的box，所以直接用set存储呗
"""
Box = namedtuple('Box', 'y x h w mark')


def get_label_file(stem, folder=None):
    filename = f'{stem}{Label.SUFFIX}'
    if folder:
        folder = Path(folder)
        return folder / filename
    else:
        return Path(filename)


class Label:
    SUFFIX = '.txt'

    def __init__(self, stem, boxes):
        # 文件名去掉后缀，用来标识每一个标签
        self.stem = stem
        # 所有框
        self.boxes = set(boxes)  # type:set[Box]

    def save_label(self, folder):
        file = get_label_file(self.stem, folder)
        with open(file, mode='w', encoding='utf-8') as fp:
            for box in self.boxes:
                line = ' '.join(map(str, box))
                fp.write(f"{line}\n")

    @classmethod
    def from_file(cls, file):
        file = Path(file)
        stem = file.stem
        boxes = set()
        with open(file, mode='r', encoding='utf-8') as fp:
            for line in fp:
                y, x, h, w, mark = line.split()
                y = int(y)
                x = int(x)
                h = int(h)
                w = int(w)
                box = Box(y, x, h, w, mark)
                if box in boxes:
                    raise ValueError(f'这个文件"{file}"里面的这个box{box}有问题， 已经存在了的，检查一下！')
                boxes.add(box)
        return cls(stem, boxes)


class Sheet:
    def __init__(self, stem, array):
        self.stem = stem
        self.array = array  # type:np.ndarray

    @classmethod
    def from_file(cls, file):
        file = Path(file)
        wb = openpyxl.open(file)
        sh = wb.worksheets[0]
        shape = (sh.max_row, sh.max_column)
        array = np.empty(shape=shape, dtype=object)
        array[::] = ''
        for row_x in range(sh.max_row):
            for col_x in range(sh.max_column):
                row = row_x + 1
                col = col_x + 1
                value = sh.cell(row, col).value
                if value:
                    array[row_x, col_x] = value
        return cls(file.stem, array)
